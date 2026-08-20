"""Generate and deliver stable, source-backed commerce output snapshots."""
from __future__ import annotations

import csv
import io
import json
from collections import defaultdict
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from app.models.commerce_output import CommerceOutput, CommerceOutputField
from app.models.conflict import DataConflict
from app.models.enrichment import EnrichmentReviewDecision, EnrichmentRun
from app.models.product import ProductAttribute, ProductRecord
from app.models.reference_data import ProductNormalizationDecision, ReferenceDataset
from app.services.reference_data_service import comparison_value
from app.services.text_normalization import normalize_product_name


CORE_FIELD_NAMES = {
    "sku_or_product_id": "SKU / Product ID",
    "product_name": "Product Name",
    "manufacturer": "Manufacturer",
    "brand": "Brand",
    "category": "Category",
}


class CommerceOutputService:
    """Materializes delivery snapshots without changing extracted source values."""

    def __init__(self, db: Session):
        self.db = db

    def _product(self, product_id: int) -> ProductRecord:
        product = (
            self.db.query(ProductRecord)
            .options(joinedload(ProductRecord.attributes).joinedload(ProductAttribute.evidence_chunks))
            .filter(ProductRecord.id == product_id)
            .first()
        )
        if not product:
            raise LookupError(f"Product {product_id} was not found.")
        return product

    def _run(self, product_id: int, enrichment_run_id: Optional[int] = None) -> EnrichmentRun:
        query = self.db.query(EnrichmentRun).filter(EnrichmentRun.product_id == product_id)
        if enrichment_run_id is not None:
            run = query.filter(EnrichmentRun.id == enrichment_run_id).first()
        else:
            run = query.order_by(EnrichmentRun.id.desc()).first()
        if not run:
            raise LookupError("No enrichment run is available for this product.")
        return run

    @staticmethod
    def _is_present(value: Any) -> bool:
        if value is None:
            return False
        text = str(value).strip()
        return bool(text) and comparison_value(text) not in {"missing", "not found", "n/a", "na", "unknown", "none", "null"}

    @staticmethod
    def _safe_json(value: Any) -> Any:
        return json.loads(json.dumps(value, default=str))

    @staticmethod
    def _evidence_from_attribute(attribute: Optional[ProductAttribute], output_attribute: Optional[dict[str, Any]] = None) -> list[dict[str, Any]]:
        if output_attribute and output_attribute.get("evidence"):
            return list(output_attribute.get("evidence") or [])
        if not attribute:
            return []
        evidence: list[dict[str, Any]] = []
        for chunk in attribute.evidence_chunks:
            evidence.append({
                "evidence_chunk_id": chunk.stable_chunk_id,
                "source_type": chunk.source_type,
                "source_identifier": chunk.source_identifier,
                "source_url": chunk.source_url,
                "page_number": chunk.page_number,
                "row_number": chunk.row_number,
                "quote": chunk.snippet_text,
            })
        if not evidence and (attribute.evidence_chunk_id or attribute.source_identifier or attribute.source_url):
            evidence.append({
                "evidence_chunk_id": attribute.evidence_chunk_id,
                "source_type": attribute.source_type,
                "source_identifier": attribute.source_identifier,
                "source_url": attribute.source_url,
                "page_number": attribute.page_number,
                "row_number": attribute.row_number,
                "quote": None,
            })
        return evidence

    def _attribute_lookup(self, product: ProductRecord) -> dict[str, ProductAttribute]:
        return {comparison_value(item.attribute_name): item for item in product.attributes}

    def _decisions(self, product_id: int) -> dict[Optional[int], list[ProductNormalizationDecision]]:
        grouped: dict[Optional[int], list[ProductNormalizationDecision]] = defaultdict(list)
        rows = self.db.query(ProductNormalizationDecision).filter(ProductNormalizationDecision.product_id == product_id).all()
        for row in rows:
            grouped[row.attribute_id].append(row)
        return grouped

    @staticmethod
    def _decision_for(field_key: str, attribute: Optional[ProductAttribute], decisions: dict[Optional[int], list[ProductNormalizationDecision]], core: bool = False) -> Optional[ProductNormalizationDecision]:
        candidates = list(decisions.get(attribute.id if attribute else None, []))
        if field_key == "manufacturer":
            preferred = [item for item in candidates if item.decision_type == "manufacturer_resolution"]
        elif field_key == "brand":
            preferred = [item for item in candidates if item.decision_type == "brand_resolution"]
        else:
            preferred = [item for item in candidates if item.decision_type in {"lov_validation", "uom_normalization"}]
        return (preferred or candidates)[-1] if (preferred or candidates) else None

    def _reference_status(self, field_key: str, value: Any, decision: Optional[ProductNormalizationDecision]) -> tuple[str, Optional[str], Optional[str]]:
        if not self._is_present(value):
            return "NOT_APPLICABLE", "No populated value was available for reference validation.", None
        if decision:
            status = decision.status or "REFERENCE_DATA_UNAVAILABLE"
            if status == "APPROVED":
                return "REFERENCE_APPROVED", decision.explanation, getattr(decision.dataset, "name", None)
            if status == "REFERENCE_DATA_UNAVAILABLE":
                return "REFERENCE_DATA_UNAVAILABLE", decision.explanation, None
            return "REFERENCE_INVALID", decision.explanation, getattr(decision.dataset, "name", None)
        # The active registry is the only authority for claiming reference compliance.
        applicable_types = {"manufacturer_brand"} if field_key in {"manufacturer", "brand"} else {"lov", "faucets_lov", "fittings_lov"} if field_key not in {"product_name", "category", "sku_or_product_id"} else set()
        if applicable_types:
            available = self.db.query(ReferenceDataset).filter(
                ReferenceDataset.is_active.is_(True),
                ReferenceDataset.status == "available",
                ReferenceDataset.dataset_type.in_(applicable_types),
            ).first()
            if not available:
                return "REFERENCE_DATA_UNAVAILABLE", "No applicable official reference dataset has been imported; no value was approved or altered.", None
        return "NOT_EVALUATED", "This field has no applicable official reference validation in the supplied data.", None

    @staticmethod
    def _character_limit(value: Any, schema_item: Optional[dict[str, Any]]) -> tuple[Optional[int], str, str]:
        if not CommerceOutputService._is_present(value):
            return None, "UNAVAILABLE", "No populated value was available for character-limit validation."
        schema_item = schema_item or {}
        # Only explicit metadata supplied by an official schema is eligible here.
        limit = schema_item.get("official_character_limit") or schema_item.get("character_limit") or schema_item.get("max_length")
        if not isinstance(limit, int) or limit <= 0:
            return None, "UNAVAILABLE", "No official character limit was supplied; this field was not scored."
        length = len(str(value))
        return limit, "PASS" if length <= limit else "FAIL", f"Official limit: {limit}; observed length: {length}."

    @staticmethod
    def _review_map(rows: list[EnrichmentReviewDecision]) -> dict[Optional[int], EnrichmentReviewDecision]:
        result: dict[Optional[int], EnrichmentReviewDecision] = {}
        for row in rows:
            result[row.attribute_id] = row
        return result

    @staticmethod
    def _conflict_map(conflicts: list[DataConflict]) -> dict[str, list[int]]:
        grouped: dict[str, list[int]] = defaultdict(list)
        for conflict in conflicts:
            grouped[comparison_value(conflict.attribute_name)].append(conflict.id)
        return grouped

    def generate(self, product_id: int, enrichment_run_id: Optional[int] = None) -> CommerceOutput:
        product = self._product(product_id)
        run = self._run(product_id, enrichment_run_id)
        output = run.output_snapshot or {}
        output_attributes = output.get("attributes") or {}
        attr_lookup = self._attribute_lookup(product)
        decisions = self._decisions(product_id)
        reviews = self.db.query(EnrichmentReviewDecision).filter(EnrichmentReviewDecision.enrichment_run_id == run.id).order_by(EnrichmentReviewDecision.id.asc()).all()
        review_by_attribute = self._review_map(reviews)
        conflicts = self.db.query(DataConflict).filter(DataConflict.product_id == product_id).order_by(DataConflict.id.asc()).all()
        conflict_by_name = self._conflict_map(conflicts)
        schema_items = {comparison_value(item.get("name")): item for item in (run.schema_snapshot or [])}

        fields: list[dict[str, Any]] = []

        def add_field(field_key: str, display_name: str, raw_value: Any, normalized_value: Any = None, unit: Optional[str] = None, confidence: Optional[float] = None, attribute: Optional[ProductAttribute] = None, output_attribute: Optional[dict[str, Any]] = None) -> None:
            value = normalized_value if self._is_present(normalized_value) else raw_value
            evidence = self._evidence_from_attribute(attribute, output_attribute)
            decision = self._decision_for(field_key, attribute, decisions, core=field_key in CORE_FIELD_NAMES)
            validation_status, validation_explanation, reference_dataset = self._reference_status(field_key, value, decision)
            limit, limit_status, limit_explanation = self._character_limit(value, schema_items.get(comparison_value(field_key)) or schema_items.get(comparison_value(display_name)))
            if limit_explanation and validation_explanation:
                validation_explanation = f"{validation_explanation} {limit_explanation}"
            elif limit_explanation:
                validation_explanation = limit_explanation
            review = review_by_attribute.get(attribute.id if attribute else None)
            conflict_ids = conflict_by_name.get(comparison_value(field_key), []) or conflict_by_name.get(comparison_value(display_name), [])
            if not self._is_present(value):
                field_status = "MISSING"
            elif conflict_ids:
                field_status = "CONFLICT"
            elif review and review.decision in {"EDIT", "REJECT", "MARK_UNRESOLVED"}:
                field_status = "REVIEW_REQUIRED"
            else:
                field_status = "PRESENT"
            fields.append({
                "field_key": field_key,
                "display_name": display_name,
                "raw_value": None if raw_value is None else str(raw_value),
                "normalized_value": None if normalized_value is None else str(normalized_value),
                "output_value": None if value is None else str(value),
                "unit": unit,
                "field_status": field_status,
                "validation_status": validation_status,
                "validation_explanation": validation_explanation,
                "reference_dataset": reference_dataset,
                "character_limit": limit,
                "character_limit_status": limit_status,
                "confidence": confidence,
                "evidence_snapshot": evidence,
                "provenance_status": "PRESENT" if evidence else "UNAVAILABLE",
                "conflict_ids": conflict_ids,
                "review_state": review.decision if review else "NOT_REVIEWED",
                "review_snapshot": ({"decision": review.decision, "value": review.reviewer_value, "reason": review.reason, "created_at": review.created_at.isoformat()} if review else None),
            })

        sku_evidence = []
        if product.sku_source_identifier or product.sku_source_url or product.sku_evidence_chunk_id:
            sku_evidence = [{"evidence_chunk_id": product.sku_evidence_chunk_id, "source_type": product.sku_source_type, "source_identifier": product.sku_source_identifier, "source_url": product.sku_source_url, "page_number": product.sku_page_number, "row_number": product.sku_row_number, "quote": None}]
        add_field("sku_or_product_id", CORE_FIELD_NAMES["sku_or_product_id"], product.sku, product.sku, None, None, attr_lookup.get("sku"), {"evidence": sku_evidence} if sku_evidence else None)
        for key, label, raw, attr_names in [
            ("product_name", CORE_FIELD_NAMES["product_name"], normalize_product_name(product.name, product.sku) or product.name, ["product_name", "name", "product title", "title"]),
            ("manufacturer", CORE_FIELD_NAMES["manufacturer"], product.manufacturer, ["manufacturer", "manufacturer name"]),
        ]:
            attribute = next((attr_lookup.get(comparison_value(name)) for name in attr_names if attr_lookup.get(comparison_value(name))), None)
            item = output_attributes.get(attribute.attribute_name if attribute else "") or output_attributes.get(key) or {}
            add_field(key, label, raw, item.get("normalized_value"), item.get("unit"), item.get("confidence"), attribute, item)
        brand_attribute = next((attr_lookup.get(key) for key in ["brand", "brand name", "manufacturer brand", "e1 brand", "unilog brand", "dib brand"] if attr_lookup.get(key)), None)
        brand_item = output_attributes.get(brand_attribute.attribute_name if brand_attribute else "") or output_attributes.get("brand") or {}
        add_field("brand", CORE_FIELD_NAMES["brand"], brand_item.get("raw_value") or brand_item.get("normalized_value"), brand_item.get("normalized_value"), brand_item.get("unit"), brand_item.get("confidence"), brand_attribute, brand_item)
        add_field("category", CORE_FIELD_NAMES["category"], run.category, run.category, None, run.category_confidence, None, None)

        core_keys = set(CORE_FIELD_NAMES)
        for name, item in output_attributes.items():
            key = comparison_value(name)
            if key in core_keys or key in {"name", "sku", "product title", "title", "manufacturer name", "brand name", "manufacturer brand", "e1 brand", "unilog brand", "dib brand"}:
                continue
            attribute = attr_lookup.get(key)
            add_field(key or name, name, item.get("raw_value"), item.get("normalized_value"), item.get("unit"), item.get("confidence"), attribute, item)

        fields_total = len(fields)
        populated = [field for field in fields if field["field_status"] != "MISSING"]
        missing = [field for field in fields if field["field_status"] == "MISSING"]
        conflict_fields = [field for field in fields if field["conflict_ids"]]
        review_fields = [field for field in fields if field["field_status"] in {"CONFLICT", "REVIEW_REQUIRED", "MISSING"} or field["validation_status"] == "REFERENCE_INVALID"]
        unavailable_reference = sum(field["validation_status"] == "REFERENCE_DATA_UNAVAILABLE" for field in fields)
        invalid_reference = sum(field["validation_status"] == "REFERENCE_INVALID" for field in fields)
        checked_limits = sum(field["character_limit_status"] in {"PASS", "FAIL"} for field in fields)
        unavailable_limits = sum(field["character_limit_status"] == "UNAVAILABLE" for field in fields)
        violations = sum(field["character_limit_status"] == "FAIL" for field in fields)
        without_provenance = sum(field["provenance_status"] == "UNAVAILABLE" for field in fields)
        reference_available = unavailable_reference < len([field for field in fields if field["validation_status"] != "NOT_APPLICABLE"]) if fields else False
        notes = ["Ground-truth accuracy is unavailable because no official expected-output dataset has been supplied."]
        if unavailable_reference:
            notes.append(f"{unavailable_reference} field(s) could not be checked because applicable official reference data was unavailable.")
        if unavailable_limits:
            notes.append(f"{unavailable_limits} field(s) have no official character limit and were not scored.")
        validation = {
            "overall_status": "READY" if not review_fields else "REVIEW_REQUIRED",
            "fields_total": fields_total,
            "fields_populated": len(populated),
            "fields_missing": len(missing),
            "fields_with_conflicts": len(conflict_fields),
            "fields_requiring_review": len(review_fields),
            "fields_without_provenance": without_provenance,
            "reference_data_available": reference_available,
            "reference_data_unavailable_fields": unavailable_reference,
            "invalid_reference_fields": invalid_reference,
            "character_limit_checked": checked_limits,
            "character_limit_unavailable": unavailable_limits,
            "character_limit_violations": violations,
            "notes": notes,
        }
        record = {
            "sku_or_product_id": next(item["output_value"] for item in fields if item["field_key"] == "sku_or_product_id"),
            "product_name": next(item["output_value"] for item in fields if item["field_key"] == "product_name"),
            "manufacturer": next(item["output_value"] for item in fields if item["field_key"] == "manufacturer"),
            "brand": next(item["output_value"] for item in fields if item["field_key"] == "brand"),
            "category": next(item["output_value"] for item in fields if item["field_key"] == "category"),
            "attributes": {item["field_key"]: item["output_value"] for item in fields if item["field_key"] not in CORE_FIELD_NAMES},
        }
        record_snapshot = {"product_id": product.id, "record": record, "raw_fields": {item["field_key"]: item["raw_value"] for item in fields}, "normalized_fields": {item["field_key"]: item["normalized_value"] for item in fields}}
        commerce = CommerceOutput(
            product_id=product.id,
            enrichment_run_id=run.id,
            output_version="1.0",
            status=validation["overall_status"],
            overall_confidence=run.overall_confidence,
            validation_summary=validation,
            record_snapshot=record_snapshot,
            source_snapshot=self._safe_json(output.get("sources") or []),
            conflict_snapshot=self._safe_json(output.get("conflicts") or [{"id": item.id, "attribute_name": item.attribute_name, "severity": item.severity, "status": item.status} for item in conflicts]),
            review_snapshot=self._safe_json([{"id": item.id, "attribute_id": item.attribute_id, "decision": item.decision, "value": item.reviewer_value, "reason": item.reason} for item in reviews]),
        )
        for field in fields:
            commerce.fields.append(CommerceOutputField(**field))
        self.db.add(commerce)
        self.db.commit()
        self.db.refresh(commerce)
        return commerce

    def latest(self, product_id: int) -> Optional[CommerceOutput]:
        return self.db.query(CommerceOutput).filter(CommerceOutput.product_id == product_id).order_by(CommerceOutput.id.desc()).first()

    def ensure_latest(self, product_id: int) -> CommerceOutput:
        latest = self.latest(product_id)
        run = self._run(product_id)
        if not latest or latest.enrichment_run_id != run.id:
            return self.generate(product_id, run.id)
        return latest

    @staticmethod
    def _output_product(output: CommerceOutput) -> dict[str, Any]:
        product = output.product
        return {"id": product.id, "sku": product.sku, "name": product.name, "manufacturer": product.manufacturer, "category": output.record_snapshot.get("record", {}).get("category")}

    def payload(self, output: CommerceOutput) -> dict[str, Any]:
        fields = []
        for field in output.fields:
            fields.append({
                "id": field.id, "field_key": field.field_key, "display_name": field.display_name,
                "raw_value": field.raw_value, "normalized_value": field.normalized_value, "output_value": field.output_value,
                "unit": field.unit, "field_status": field.field_status, "validation_status": field.validation_status,
                "validation_explanation": field.validation_explanation, "reference_dataset": field.reference_dataset,
                "character_limit": field.character_limit, "character_limit_status": field.character_limit_status,
                "confidence": field.confidence, "evidence": field.evidence_snapshot or [], "provenance_status": field.provenance_status, "conflict_ids": field.conflict_ids or [],
                "review_state": field.review_state, "review": field.review_snapshot,
            })
        return {
            "id": output.id, "product_id": output.product_id, "enrichment_run_id": output.enrichment_run_id,
            "output_version": output.output_version, "status": output.status, "overall_confidence": output.overall_confidence,
            "generated_at": output.generated_at, "product": self._output_product(output),
            "record": output.record_snapshot.get("record", {}), "fields": fields,
            "validation": output.validation_summary or {}, "sources": output.source_snapshot or [],
            "conflicts": output.conflict_snapshot or [], "reviews": output.review_snapshot or [],
            "ground_truth_accuracy": "UNAVAILABLE",
        }

    def export(self, output: CommerceOutput, fmt: str) -> tuple[bytes, str, str]:
        payload = self.payload(output)
        if fmt == "json":
            return json.dumps(payload, default=str, indent=2).encode("utf-8"), "application/json", f"commerce-output-{output.product_id}.json"
        rows = []
        for field in payload["fields"]:
            rows.append({
                "product_id": output.product_id, "field_key": field["field_key"], "display_name": field["display_name"],
                "raw_value": field["raw_value"], "normalized_value": field["normalized_value"], "output_value": field["output_value"],
                "unit": field["unit"], "field_status": field["field_status"], "validation_status": field["validation_status"],
                "confidence": field["confidence"], "character_limit_status": field["character_limit_status"],
                "review_state": field["review_state"], "evidence_count": len(field["evidence"]), "conflict_ids": ",".join(map(str, field["conflict_ids"])),
            })
        headers = list(rows[0].keys()) if rows else ["product_id", "field_key", "output_value"]
        if fmt == "csv":
            stream = io.StringIO()
            writer = csv.DictWriter(stream, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            return stream.getvalue().encode("utf-8"), "text/csv", f"commerce-output-{output.product_id}.csv"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fields"
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        product_sheet = workbook.create_sheet("Product")
        product_sheet.append(["key", "value"])
        for key, value in payload["record"].items():
            product_sheet.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])
        validation_sheet = workbook.create_sheet("Validation")
        validation_sheet.append(["key", "value"])
        for key, value in payload["validation"].items():
            validation_sheet.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"commerce-output-{output.product_id}.xlsx"
